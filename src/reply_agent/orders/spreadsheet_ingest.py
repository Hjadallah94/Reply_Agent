"""Parses a seller's order-tracking sheet (Doc 2 Section 2.6: "many sellers track orders in a
Google Sheet or Excel file — support a simple, low-friction sync"). A single-sheet workbook,
synced separately from the catalog (scripts/ingest_catalog.py) since orders change on a much
faster cadence than the product/policy/FAQ knowledge base.

Expected workbook: one .xlsx file with an "Orders" sheet — order_reference, customer_phone,
status (all required); customer_name, items_summary, order_date (optional, "YYYY-MM-DD" or an
Excel date cell). A bad individual row is skipped and reported, not fatal to the rest of the
sheet; a missing sheet or required column is a hard error.
"""

from datetime import UTC, date, datetime

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from pydantic import ValidationError

from reply_agent.orders.phone import normalize_phone
from reply_agent.orders.schema import OrderRecord


class OrderWorkbookParseError(ValueError):
    """The workbook is missing the Orders sheet, or that sheet is missing a required column."""


def _read_rows(sheet: Worksheet) -> tuple[list[str], list[dict[str, object]]]:
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return [], []
    headers = [str(h).strip().lower() if h is not None else "" for h in rows[0]]
    data = [
        {headers[i]: row[i] for i in range(len(headers)) if i < len(row)}
        for row in rows[1:]
        if not all(cell is None for cell in row)
    ]
    return headers, data


def _parse_order_date(cell: object) -> datetime | None:
    if cell is None or str(cell).strip() == "":
        return None
    if isinstance(cell, datetime):
        return cell if cell.tzinfo else cell.replace(tzinfo=UTC)
    if isinstance(cell, date):
        return datetime(cell.year, cell.month, cell.day, tzinfo=UTC)
    return datetime.strptime(str(cell).strip(), "%Y-%m-%d").replace(tzinfo=UTC)


def parse_orders_workbook(file) -> tuple[list[OrderRecord], list[str]]:
    """`file` is anything openpyxl's load_workbook accepts: a path (str/Path) or a file-like
    object such as io.BytesIO."""
    wb = load_workbook(file, read_only=True, data_only=True)
    try:
        if "Orders" not in wb.sheetnames:
            raise OrderWorkbookParseError("Workbook is missing an 'Orders' sheet")

        headers, rows = _read_rows(wb["Orders"])
        required = ["order_reference", "customer_phone", "status"]
        missing = [c for c in required if c not in headers]
        if missing:
            raise OrderWorkbookParseError(
                f"'Orders' sheet is missing required column(s): {', '.join(missing)}"
            )

        records: list[OrderRecord] = []
        issues: list[str] = []
        for i, row in enumerate(rows, start=2):
            try:
                order_reference = str(row.get("order_reference") or "").strip()
                phone_raw = row.get("customer_phone")
                status = str(row.get("status") or "").strip()
                if not order_reference or not phone_raw or not status:
                    raise ValueError("order_reference, customer_phone, and status are all required")
                name = row.get("customer_name")
                items = row.get("items_summary")
                records.append(
                    OrderRecord(
                        order_reference=order_reference,
                        customer_phone=normalize_phone(str(phone_raw)),
                        customer_name=str(name).strip() if name else None,
                        status=status,
                        items_summary=str(items).strip() if items else None,
                        order_date=_parse_order_date(row.get("order_date")),
                    )
                )
            except (ValidationError, ValueError, TypeError) as exc:
                issues.append(f"Orders row {i}: skipped — {exc}")
    finally:
        wb.close()

    return records, issues
