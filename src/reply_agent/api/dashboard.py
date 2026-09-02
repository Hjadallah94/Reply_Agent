"""Owner-facing dashboard (Doc 3 Phase 3): view conversations, approve/edit/send escalated
replies. Server-rendered with Jinja2 rather than a separate frontend build. Every route below
is gated by auth/dependencies.py's require_business_access — a logged-in user only ever sees
their own business, never anyone else's — and reads/writes through db/tenant_session.py, which
enforces that same boundary again at the database level (row-level security, not just app code).
"""

import uuid
from datetime import UTC, datetime, timedelta
from functools import partial
from io import BytesIO
from pathlib import Path
from zoneinfo import ZoneInfo

from fastapi import APIRouter, Depends, Form, HTTPException, Request
from fastapi.responses import RedirectResponse, StreamingResponse
from fastapi.templating import Jinja2Templates
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from pydantic import BaseModel, ValidationError
from sqlalchemy import func, select
from sqlalchemy.orm import selectinload

from reply_agent.auth.dependencies import (
    ensure_business_access,
    get_current_user,
    require_business_access,
)
from reply_agent.billing.usage import get_or_create_subscription, usage_summary
from reply_agent.config import get_settings
from reply_agent.db.models import (
    ApprovalRequest,
    ApprovalRequestStatus,
    Business,
    Conversation,
    ConversationStatus,
    Customer,
    CustomRule,
    Escalation,
    EscalationStatus,
    KnowledgeDocType,
    KnowledgeDocument,
    Message,
    MessageDirection,
    Order,
    PushSubscription,
)
from reply_agent.db.tenant_session import tenant_session
from reply_agent.graph.nodes.request_owner_approval import AUTO_APPROVAL_RESOLVED_BY
from reply_agent.graph.nodes.send_reply import send_reply
from reply_agent.graph.risk_rules import (
    DEFAULT_SENSITIVITY,
    RISK_INTENT_LABELS,
    SENSITIVITY_THRESHOLDS,
)
from reply_agent.i18n import SUPPORTED_LANGUAGES, get_lang, t, t_status
from reply_agent.knowledge.catalog import (
    create_product,
    create_promotion,
    update_product,
    update_promotion,
)
from reply_agent.knowledge.corrections import record_owner_correction
from reply_agent.knowledge.schema import Product, Promotion
from reply_agent.knowledge.spreadsheet_ingest import parse_variants
from reply_agent.theming import THEME_OPTIONS, get_theme

router = APIRouter(tags=["dashboard"])
templates = Jinja2Templates(directory=str(Path(__file__).resolve().parent.parent / "templates"))

# Same convention as graph/nodes/estimate_delivery.py — business-facing times in this project
# are Amman-local, not UTC; the catalog forms' datetime-local inputs are entered in this zone.
AMMAN_TZ = ZoneInfo("Asia/Amman")


def _render(request: Request, name: str, **context):
    """Every dashboard template render goes through here (Doc 3 Phase 6.6) so lang/t/t_status
    can never be forgotten on one route — a plain templates.TemplateResponse(...) call would
    silently render English-only chrome regardless of the viewer's language preference.
    """
    lang = get_lang(request)
    context.setdefault("lang", lang)
    context.setdefault("t", partial(t, lang))
    context.setdefault("t_status", partial(t_status, lang))
    context.setdefault("theme", get_theme(request))
    context.setdefault("amman_tz", AMMAN_TZ)
    context.setdefault("vapid_public_key", get_settings().vapid_public_key)
    return templates.TemplateResponse(request, name, context)


class _PushSubscriptionKeys(BaseModel):
    p256dh: str
    auth: str


class _PushSubscribePayload(BaseModel):
    business_id: uuid.UUID
    endpoint: str
    keys: _PushSubscriptionKeys


class _PushUnsubscribePayload(BaseModel):
    business_id: uuid.UUID
    endpoint: str


@router.post("/push-subscribe")
async def push_subscribe(request: Request, payload: _PushSubscribePayload) -> dict:
    await ensure_business_access(request, payload.business_id)
    user = await get_current_user(request)

    async with tenant_session(payload.business_id) as session:
        # Upsert on endpoint, not a duplicate — the same browser re-subscribing (e.g. after
        # clearing site data then re-granting permission) naturally reuses or gets issued a
        # fresh endpoint from the push service; either way this keeps exactly one row per
        # endpoint rather than accumulating dead duplicates.
        existing = await session.scalar(
            select(PushSubscription).where(PushSubscription.endpoint == payload.endpoint)
        )
        if existing is not None:
            existing.business_id = payload.business_id
            existing.user_id = user.id
            existing.p256dh_key = payload.keys.p256dh
            existing.auth_key = payload.keys.auth
        else:
            session.add(
                PushSubscription(
                    business_id=payload.business_id,
                    user_id=user.id,
                    endpoint=payload.endpoint,
                    p256dh_key=payload.keys.p256dh,
                    auth_key=payload.keys.auth,
                )
            )

    return {"status": "ok"}


@router.post("/push-unsubscribe")
async def push_unsubscribe(request: Request, payload: _PushUnsubscribePayload) -> dict:
    await ensure_business_access(request, payload.business_id)

    async with tenant_session(payload.business_id) as session:
        existing = await session.scalar(
            select(PushSubscription).where(PushSubscription.endpoint == payload.endpoint)
        )
        if existing is not None:
            await session.delete(existing)

    return {"status": "ok"}


@router.get("/dashboard")
async def dashboard_redirect(request: Request):
    # No multi-business list — a user has exactly one business, straight there or to /login.
    user = await get_current_user(request)
    return RedirectResponse(url=f"/businesses/{user.business_id}/dashboard", status_code=303)


@router.post("/set-language")
async def set_language(request: Request, lang: str = Form(...), next: str = Form("/dashboard")):
    if lang not in SUPPORTED_LANGUAGES:
        raise HTTPException(status_code=400, detail="Unsupported language")
    # Open-redirect guard: only ever redirect to a same-origin relative path, never an
    # absolute or protocol-relative URL a malicious `next` value could point elsewhere.
    safe_next = next if next.startswith("/") and not next.startswith("//") else "/dashboard"
    request.session["lang"] = lang
    return RedirectResponse(url=safe_next, status_code=303)


@router.post("/set-theme")
async def set_theme(request: Request, theme: str = Form(...), next: str = Form("/dashboard")):
    if theme not in THEME_OPTIONS:
        raise HTTPException(status_code=400, detail="Unsupported theme")
    # Same open-redirect guard as /set-language above.
    safe_next = next if next.startswith("/") and not next.startswith("//") else "/dashboard"
    request.session["theme"] = theme
    return RedirectResponse(url=safe_next, status_code=303)


@router.get("/businesses/{business_id}/dashboard")
async def business_dashboard(
    request: Request, business: Business = Depends(require_business_access)
):
    async with tenant_session(business.id) as session:
        subscription = await get_or_create_subscription(session, business)
        usage = usage_summary(subscription)

        pending = (
            await session.scalars(
                select(Escalation)
                .join(Conversation)
                .where(
                    Conversation.business_id == business.id,
                    Escalation.status == EscalationStatus.pending,
                )
                .options(selectinload(Escalation.conversation).selectinload(Conversation.customer))
                .order_by(Escalation.created_at)
            )
        ).all()

        pending_rows = [
            {
                "id": e.id,
                "customer_handle": e.conversation.customer.channel_handle,
                "channel": e.conversation.channel.value,
                "reason": e.reason,
            }
            for e in pending
        ]

        pending_approval_records = (
            await session.scalars(
                select(ApprovalRequest)
                .join(Conversation)
                .where(
                    Conversation.business_id == business.id,
                    ApprovalRequest.status == ApprovalRequestStatus.pending,
                )
                .options(
                    selectinload(ApprovalRequest.conversation).selectinload(Conversation.customer)
                )
                .order_by(ApprovalRequest.created_at)
            )
        ).all()

        pending_approval_rows = [
            {
                "id": a.id,
                "customer_handle": a.conversation.customer.channel_handle,
                "channel": a.conversation.channel.value,
                "reasoning": a.reasoning,
            }
            for a in pending_approval_records
        ]

        # Adaptive autonomy's explainability surface (Doc 2 Section 9.4): read-only visibility
        # into what the system has started sending on its own, so the owner can always see why —
        # not a UI they act on, unlike the two sections above.
        recent_auto_approvals = (
            await session.scalars(
                select(ApprovalRequest)
                .join(Conversation)
                .where(
                    Conversation.business_id == business.id,
                    ApprovalRequest.resolved_by == AUTO_APPROVAL_RESOLVED_BY,
                    ApprovalRequest.resolution_time >= datetime.now(UTC) - timedelta(days=7),
                )
                .options(
                    selectinload(ApprovalRequest.conversation).selectinload(Conversation.customer)
                )
                .order_by(ApprovalRequest.resolution_time.desc())
            )
        ).all()

        recent_auto_approval_rows = [
            {
                "customer_handle": a.conversation.customer.channel_handle,
                "channel": a.conversation.channel.value,
                "estimated_window": a.estimated_window,
                "resolution_time": a.resolution_time.strftime("%Y-%m-%d %H:%M"),
            }
            for a in recent_auto_approvals
        ]

        conversations = (
            await session.scalars(
                select(Conversation)
                .where(Conversation.business_id == business.id)
                .options(selectinload(Conversation.customer), selectinload(Conversation.messages))
                .order_by(Conversation.updated_at.desc())
                .limit(30)
            )
        ).all()

        conversation_rows = [
            {
                "id": c.id,
                "customer_handle": c.customer.channel_handle,
                "status": c.status.value,
                "last_message_text": c.messages[-1].text if c.messages else None,
            }
            for c in conversations
        ]

    return _render(
        request,
        "dashboard.html",
        business=business,
        usage=usage,
        pending_escalations=pending_rows,
        pending_approvals=pending_approval_rows,
        recent_auto_approvals=recent_auto_approval_rows,
        conversations=conversation_rows,
    )


@router.post("/businesses/{business_id}/away-mode")
async def set_away_mode(
    business_id: uuid.UUID,
    is_away: bool = Form(False),
    away_message: str = Form(""),
    business: Business = Depends(require_business_access),
):
    async with tenant_session(business_id) as session:
        db_business = await session.get(Business, business_id)
        db_business.is_away = is_away
        db_business.away_message = away_message.strip() or None

    return RedirectResponse(url=f"/businesses/{business.id}/dashboard", status_code=303)


CONVERSATIONS_PAGE_SIZE = 50


@router.get("/businesses/{business_id}/dashboard/conversations")
async def conversations_list(
    request: Request,
    page: int = 1,
    business: Business = Depends(require_business_access),
):
    # Doc 3 roadmap (partner meeting 2026-09-01): the dashboard's own "Recent conversations"
    # section (business_dashboard above) is a fixed top-30 preview — this is the actual
    # "see all the interactions that happened" surface the meeting note asked for.
    page = max(1, page)
    async with tenant_session(business.id) as session:
        total = await session.scalar(
            select(func.count())
            .select_from(Conversation)
            .where(Conversation.business_id == business.id)
        )
        conversations = (
            await session.scalars(
                select(Conversation)
                .where(Conversation.business_id == business.id)
                .options(selectinload(Conversation.customer), selectinload(Conversation.messages))
                .order_by(Conversation.updated_at.desc())
                .limit(CONVERSATIONS_PAGE_SIZE)
                .offset((page - 1) * CONVERSATIONS_PAGE_SIZE)
            )
        ).all()

    conversation_rows = [
        {
            "id": c.id,
            "customer_handle": c.customer.channel_handle,
            "channel": c.channel.value,
            "status": c.status.value,
            "last_message_text": c.messages[-1].text if c.messages else None,
        }
        for c in conversations
    ]
    total_pages = max(1, -(-(total or 0) // CONVERSATIONS_PAGE_SIZE))
    return _render(
        request,
        "conversations_list.html",
        business=business,
        conversations=conversation_rows,
        page=page,
        total_pages=total_pages,
    )


@router.get("/businesses/{business_id}/dashboard/conversations/{conversation_id}")
async def conversation_detail(
    request: Request,
    conversation_id: uuid.UUID,
    business: Business = Depends(require_business_access),
):
    async with tenant_session(business.id) as session:
        conversation = await session.scalar(
            select(Conversation)
            .where(Conversation.id == conversation_id)
            .options(selectinload(Conversation.customer), selectinload(Conversation.messages))
        )
        if conversation is None or conversation.business_id != business.id:
            raise HTTPException(status_code=404, detail="Conversation not found")

        # A pending Escalation/ApprovalRequest already has its own dedicated resolve/approve
        # flow (with correction-tracking and adaptive-autonomy bookkeeping this general send
        # form doesn't do) — point there instead of inviting a parallel free-form reply that
        # would leave that row stuck "pending" forever.
        pending_escalation = await session.scalar(
            select(Escalation).where(
                Escalation.conversation_id == conversation.id,
                Escalation.status == EscalationStatus.pending,
            )
        )
        pending_approval = await session.scalar(
            select(ApprovalRequest).where(
                ApprovalRequest.conversation_id == conversation.id,
                ApprovalRequest.status == ApprovalRequestStatus.pending,
            )
        )

        messages = [
            {
                "direction": m.direction.value,
                "text": m.text,
                "created_at": m.created_at.strftime("%Y-%m-%d %H:%M"),
            }
            for m in conversation.messages
        ]

    return _render(
        request,
        "conversation.html",
        business=business,
        conversation=conversation,
        customer_handle=conversation.customer.channel_handle,
        channel=conversation.channel.value,
        messages=messages,
        pending_escalation_id=pending_escalation.id if pending_escalation else None,
        pending_approval_id=pending_approval.id if pending_approval else None,
    )


@router.post("/businesses/{business_id}/dashboard/conversations/{conversation_id}/send")
async def send_conversation_message(
    conversation_id: uuid.UUID,
    reply_text: str = Form(...),
    business: Business = Depends(require_business_access),
):
    # Same CRLF-normalization fix as resolve_escalation below — a <textarea> submission always
    # normalizes line breaks to \r\n regardless of what was typed.
    reply_text = reply_text.strip().replace("\r\n", "\n")
    if not reply_text:
        raise HTTPException(status_code=400, detail="Reply text is required")

    async with tenant_session(business.id) as session:
        conversation = await session.get(Conversation, conversation_id)
        if conversation is None or conversation.business_id != business.id:
            raise HTTPException(status_code=404, detail="Conversation not found")

        # Reuses the same channel-dispatch code a real auto-send would go through, rather than
        # duplicating the WhatsApp/Instagram/Messenger match statement here.
        await send_reply(
            {
                "channel": conversation.channel.value,
                "business_id": str(conversation.business_id),
                "thread_id": conversation.thread_id,
                "draft_reply": {"text": reply_text},
            }
        )

        session.add(
            Message(
                conversation_id=conversation.id,
                direction=MessageDirection.outbound,
                text=reply_text,
                model_used="owner",
            )
        )

    return RedirectResponse(
        url=f"/businesses/{business.id}/dashboard/conversations/{conversation_id}",
        status_code=303,
    )


@router.get("/businesses/{business_id}/dashboard/rules")
async def rules_page(request: Request, business: Business = Depends(require_business_access)):
    async with tenant_session(business.id) as session:
        custom_rules = (
            await session.scalars(
                select(CustomRule)
                .where(CustomRule.business_id == business.id)
                .order_by(CustomRule.created_at.desc())
            )
        ).all()

    escalation_rules = business.escalation_rules or {}
    excluded_locations = "\n".join((business.delivery_rules or {}).get("excluded_locations", []))

    return _render(
        request,
        "rules.html",
        business=business,
        risk_categories=set(escalation_rules.get("risk_categories", RISK_INTENT_LABELS)),
        sensitivity=escalation_rules.get("sensitivity", DEFAULT_SENSITIVITY),
        excluded_locations=excluded_locations,
        custom_rules=custom_rules,
    )


@router.post("/businesses/{business_id}/dashboard/rules/autonomy")
async def save_autonomy_rules(
    business_id: uuid.UUID,
    price_negotiation: bool = Form(False),
    refund_or_complaint: bool = Form(False),
    competitor_mention: bool = Form(False),
    legal_threat: bool = Form(False),
    sensitivity: str = Form(DEFAULT_SENSITIVITY),
    business: Business = Depends(require_business_access),
):
    if sensitivity not in SENSITIVITY_THRESHOLDS:
        raise HTTPException(status_code=400, detail="Invalid sensitivity")

    risk_categories = [
        label
        for label, checked in [
            ("price_negotiation", price_negotiation),
            ("refund_or_complaint", refund_or_complaint),
            ("competitor_mention", competitor_mention),
            ("legal_threat", legal_threat),
        ]
        if checked
    ]

    async with tenant_session(business_id) as session:
        db_business = await session.get(Business, business_id)
        db_business.escalation_rules = {
            **(db_business.escalation_rules or {}),
            "risk_categories": risk_categories,
            "sensitivity": sensitivity,
        }

    return RedirectResponse(url=f"/businesses/{business.id}/dashboard/rules", status_code=303)


@router.post("/businesses/{business_id}/dashboard/rules/delivery-restrictions")
async def save_delivery_restrictions(
    business_id: uuid.UUID,
    excluded_locations: str = Form(""),
    business: Business = Depends(require_business_access),
):
    locations = [line.strip() for line in excluded_locations.splitlines() if line.strip()]

    async with tenant_session(business_id) as session:
        db_business = await session.get(Business, business_id)
        db_business.delivery_rules = {
            **(db_business.delivery_rules or {}),
            "excluded_locations": locations,
        }

    return RedirectResponse(url=f"/businesses/{business.id}/dashboard/rules", status_code=303)


@router.post("/businesses/{business_id}/dashboard/rules/custom")
async def submit_custom_rule(
    business_id: uuid.UUID,
    rule_text: str = Form(...),
    business: Business = Depends(require_business_access),
):
    rule_text = rule_text.strip()
    if not rule_text:
        raise HTTPException(status_code=400, detail="Rule text is required")

    async with tenant_session(business_id) as session:
        session.add(CustomRule(business_id=business_id, rule_text=rule_text))

    return RedirectResponse(url=f"/businesses/{business.id}/dashboard/rules", status_code=303)


@router.get("/businesses/{business_id}/dashboard/export")
async def export_conversations(business: Business = Depends(require_business_access)):
    async with tenant_session(business.id) as session:
        rows = (
            await session.execute(
                select(Message, Conversation, Customer)
                .join(Conversation, Message.conversation_id == Conversation.id)
                .join(Customer, Conversation.customer_id == Customer.id)
                .where(Conversation.business_id == business.id)
                .order_by(Customer.channel_handle, Message.created_at)
            )
        ).all()

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Conversations"
    headers = [
        "Customer",
        "Channel",
        "From",
        "Message",
        "Intent",
        "Handled by",
        "Conversation status",
        "Sent at (UTC)",
    ]
    sheet.append(headers)
    for message, conversation, customer in rows:
        sheet.append(
            [
                customer.channel_handle,
                conversation.channel.value,
                "Customer" if message.direction == MessageDirection.inbound else "Agent",
                message.text,
                message.intent_label or "",
                message.model_used or "",
                conversation.status.value,
                message.created_at.strftime("%Y-%m-%d %H:%M"),
            ]
        )

    for i, width in enumerate([16, 11, 10, 60, 22, 16, 16, 16], start=1):
        sheet.column_dimensions[get_column_letter(i)].width = width
    sheet.freeze_panes = "A2"

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)

    safe_name = "".join(c if c.isalnum() else "_" for c in business.name)
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{safe_name}_conversations.xlsx"'},
    )


@router.get("/businesses/{business_id}/dashboard/escalations/{escalation_id}")
async def escalation_detail(
    request: Request,
    escalation_id: uuid.UUID,
    business: Business = Depends(require_business_access),
):
    async with tenant_session(business.id) as session:
        escalation = await session.scalar(
            select(Escalation)
            .where(Escalation.id == escalation_id)
            .options(
                selectinload(Escalation.conversation).selectinload(Conversation.customer),
                selectinload(Escalation.conversation).selectinload(Conversation.messages),
            )
        )
        if escalation is None or escalation.conversation.business_id != business.id:
            raise HTTPException(status_code=404, detail="Escalation not found")

        conversation = escalation.conversation
        messages = [
            {
                "direction": m.direction.value,
                "text": m.text,
                "created_at": m.created_at.strftime("%Y-%m-%d %H:%M"),
            }
            for m in conversation.messages
        ]

    return _render(
        request,
        "escalation.html",
        business=business,
        escalation=escalation,
        customer_handle=conversation.customer.channel_handle,
        channel=conversation.channel.value,
        messages=messages,
    )


@router.post("/businesses/{business_id}/dashboard/escalations/{escalation_id}/resolve")
async def resolve_escalation(
    escalation_id: uuid.UUID,
    reply_text: str = Form(...),
    business: Business = Depends(require_business_access),
):
    # A <textarea> form submission always normalizes line breaks to \r\n (CRLF) per the HTML
    # spec, but the LLM's own drafted_reply uses plain \n — normalize before comparing, or an
    # unedited approval/resolve would spuriously look "changed" every single time, breaking the
    # adaptive-autonomy streak (graph/nodes/request_owner_approval.py) and falsely recording an
    # owner-correction that's actually the model's own words. Found live during Phase 6d
    # verification: a real dashboard approve, no edits made, still failed `== drafted_reply`.
    reply_text = reply_text.strip().replace("\r\n", "\n")
    if not reply_text:
        raise HTTPException(status_code=400, detail="Reply text is required")

    async with tenant_session(business.id) as session:
        escalation = await session.scalar(
            select(Escalation)
            .where(Escalation.id == escalation_id)
            .options(selectinload(Escalation.conversation))
        )
        if escalation is None or escalation.conversation.business_id != business.id:
            raise HTTPException(status_code=404, detail="Escalation not found")
        if escalation.status != EscalationStatus.pending:
            raise HTTPException(status_code=409, detail="Escalation already resolved")

        conversation = escalation.conversation

        # Reuses the same channel-dispatch code a real auto-send would go through, rather
        # than duplicating the WhatsApp/Instagram/Messenger match statement here.
        await send_reply(
            {
                "channel": conversation.channel.value,
                "business_id": str(conversation.business_id),
                "thread_id": conversation.thread_id,
                "draft_reply": {"text": reply_text},
            }
        )

        # update_memory only logs an outbound message when route == "send" (graph/nodes/
        # update_memory.py) — escalated turns skip that, so this is the first record of it.
        session.add(
            Message(
                conversation_id=conversation.id,
                direction=MessageDirection.outbound,
                text=reply_text,
                model_used="owner",
            )
        )

        # Owner-correction feedback loop (Doc 1 Section 7) — only when the owner actually sent
        # something different from the agent's own draft (including drafting nothing at all, a
        # capability-gap escalation): approving a draft unchanged isn't a correction, it's
        # confirmation the agent already had it right, and doesn't need to teach it anything.
        if reply_text != (escalation.drafted_reply or ""):
            customer_message = await session.scalar(
                select(Message)
                .where(
                    Message.conversation_id == conversation.id,
                    Message.direction == MessageDirection.inbound,
                )
                .order_by(Message.created_at.desc())
                .limit(1)
            )
            if customer_message is not None:
                await record_owner_correction(
                    session,
                    business_id=business.id,
                    customer_message=customer_message.text,
                    corrected_reply=reply_text,
                    escalation_id=escalation.id,
                )

        escalation.status = EscalationStatus.resolved
        escalation.resolved_by = "owner"
        escalation.resolution_text = reply_text
        escalation.resolution_time = datetime.now(UTC)
        conversation.status = ConversationStatus.auto

    return RedirectResponse(url=f"/businesses/{business.id}/dashboard", status_code=303)


@router.get("/businesses/{business_id}/dashboard/approvals/{approval_id}")
async def approval_detail(
    request: Request,
    approval_id: uuid.UUID,
    business: Business = Depends(require_business_access),
):
    async with tenant_session(business.id) as session:
        approval = await session.scalar(
            select(ApprovalRequest)
            .where(ApprovalRequest.id == approval_id)
            .options(
                selectinload(ApprovalRequest.conversation).selectinload(Conversation.customer),
                selectinload(ApprovalRequest.conversation).selectinload(Conversation.messages),
            )
        )
        if approval is None or approval.conversation.business_id != business.id:
            raise HTTPException(status_code=404, detail="Approval request not found")

        conversation = approval.conversation
        messages = [
            {
                "direction": m.direction.value,
                "text": m.text,
                "created_at": m.created_at.strftime("%Y-%m-%d %H:%M"),
            }
            for m in conversation.messages
        ]

    lang = get_lang(request)
    return _render(
        request,
        "approval.html",
        business=business,
        approval=approval,
        customer_handle=conversation.customer.channel_handle,
        channel=conversation.channel.value,
        messages=messages,
        reject_default=t(lang, "approval.reject_default_text"),
    )


@router.post("/businesses/{business_id}/dashboard/approvals/{approval_id}/approve")
async def approve_approval(
    approval_id: uuid.UUID,
    reply_text: str = Form(...),
    business: Business = Depends(require_business_access),
):
    # A <textarea> form submission always normalizes line breaks to \r\n (CRLF) per the HTML
    # spec, but the LLM's own drafted_reply uses plain \n — normalize before comparing, or an
    # unedited approval/resolve would spuriously look "changed" every single time, breaking the
    # adaptive-autonomy streak (graph/nodes/request_owner_approval.py) and falsely recording an
    # owner-correction that's actually the model's own words. Found live during Phase 6d
    # verification: a real dashboard approve, no edits made, still failed `== drafted_reply`.
    reply_text = reply_text.strip().replace("\r\n", "\n")
    if not reply_text:
        raise HTTPException(status_code=400, detail="Reply text is required")

    async with tenant_session(business.id) as session:
        approval = await session.scalar(
            select(ApprovalRequest)
            .where(ApprovalRequest.id == approval_id)
            .options(selectinload(ApprovalRequest.conversation))
        )
        if approval is None or approval.conversation.business_id != business.id:
            raise HTTPException(status_code=404, detail="Approval request not found")
        if approval.status != ApprovalRequestStatus.pending:
            raise HTTPException(status_code=409, detail="Approval request already resolved")

        conversation = approval.conversation

        await send_reply(
            {
                "channel": conversation.channel.value,
                "business_id": str(conversation.business_id),
                "thread_id": conversation.thread_id,
                "draft_reply": {"text": reply_text},
            }
        )

        session.add(
            Message(
                conversation_id=conversation.id,
                direction=MessageDirection.outbound,
                text=reply_text,
                model_used="owner",
            )
        )

        # Same owner-correction gate as resolve_escalation — approving a draft unchanged isn't
        # a correction, editing it before sending is.
        if reply_text != (approval.drafted_reply or ""):
            customer_message = await session.scalar(
                select(Message)
                .where(
                    Message.conversation_id == conversation.id,
                    Message.direction == MessageDirection.inbound,
                )
                .order_by(Message.created_at.desc())
                .limit(1)
            )
            if customer_message is not None:
                await record_owner_correction(
                    session,
                    business_id=business.id,
                    customer_message=customer_message.text,
                    corrected_reply=reply_text,
                    approval_id=approval.id,
                )

        approval.status = ApprovalRequestStatus.approved
        approval.resolved_by = "owner"
        approval.resolution_time = datetime.now(UTC)
        # Adaptive autonomy's training signal (Doc 2 Section 9.4, graph/nodes/
        # request_owner_approval.py's _matches_learned_pattern) — only an unedited approval
        # counts toward earning auto-approval for this (business, estimated_window) pattern.
        approval.sent_unchanged = reply_text == (approval.drafted_reply or "")
        conversation.status = ConversationStatus.auto

    return RedirectResponse(url=f"/businesses/{business.id}/dashboard", status_code=303)


def _document_or_404(
    document: KnowledgeDocument | None, business_id: uuid.UUID, expected_type: KnowledgeDocType
) -> KnowledgeDocument:
    if document is None or document.business_id != business_id or document.type != expected_type:
        raise HTTPException(status_code=404, detail="Not found")
    return document


def _parse_amman_datetime(value: str, *, field_name: str) -> datetime:
    try:
        naive = datetime.strptime(value, "%Y-%m-%dT%H:%M")
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=f"Invalid {field_name}") from exc
    return naive.replace(tzinfo=AMMAN_TZ)


@router.get("/businesses/{business_id}/dashboard/catalog")
async def catalog_list(request: Request, business: Business = Depends(require_business_access)):
    async with tenant_session(business.id) as session:
        products = (
            await session.scalars(
                select(KnowledgeDocument)
                .where(
                    KnowledgeDocument.business_id == business.id,
                    KnowledgeDocument.type == KnowledgeDocType.product,
                )
                .order_by(KnowledgeDocument.structured_data["name"].astext)
            )
        ).all()

        promotions = (
            await session.scalars(
                select(KnowledgeDocument)
                .where(
                    KnowledgeDocument.business_id == business.id,
                    KnowledgeDocument.type == KnowledgeDocType.promotion,
                )
                .order_by(KnowledgeDocument.active_until.desc())
            )
        ).all()

    now = datetime.now(UTC)
    return _render(
        request,
        "catalog.html",
        business=business,
        products=products,
        promotions=promotions,
        now=now,
    )


@router.get("/businesses/{business_id}/dashboard/catalog/products/new")
async def new_product_form(request: Request, business: Business = Depends(require_business_access)):
    return _render(
        request,
        "product_form.html",
        business=business,
        document=None,
        product=None,
        variants_text="",
    )


@router.post("/businesses/{business_id}/dashboard/catalog/products/new")
async def create_product_route(
    name: str = Form(...),
    description: str = Form(""),
    price_jod: str = Form(...),
    stock_status: str = Form("in_stock"),
    variants: str = Form(""),
    business: Business = Depends(require_business_access),
):
    try:
        product = Product(
            name=name.strip(),
            description=description.strip(),
            price_jod=float(price_jod),
            stock_status=stock_status.strip() or "in_stock",
            variants=parse_variants(variants),
        )
    except (ValidationError, ValueError) as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc

    async with tenant_session(business.id) as session:
        await create_product(session, business.id, product)

    return RedirectResponse(url=f"/businesses/{business.id}/dashboard/catalog", status_code=303)


@router.get("/businesses/{business_id}/dashboard/catalog/products/{document_id}/edit")
async def edit_product_form(
    request: Request,
    document_id: uuid.UUID,
    business: Business = Depends(require_business_access),
):
    async with tenant_session(business.id) as session:
        document = _document_or_404(
            await session.get(KnowledgeDocument, document_id), business.id, KnowledgeDocType.product
        )

    variants_text = "; ".join(
        f"{v['label']}:{v['stock_status']}" for v in document.structured_data.get("variants", [])
    )
    return _render(
        request,
        "product_form.html",
        business=business,
        document=document,
        product=document.structured_data,
        variants_text=variants_text,
    )


@router.post("/businesses/{business_id}/dashboard/catalog/products/{document_id}/edit")
async def update_product_route(
    document_id: uuid.UUID,
    name: str = Form(...),
    description: str = Form(""),
    price_jod: str = Form(...),
    stock_status: str = Form("in_stock"),
    variants: str = Form(""),
    business: Business = Depends(require_business_access),
):
    try:
        product = Product(
            name=name.strip(),
            description=description.strip(),
            price_jod=float(price_jod),
            stock_status=stock_status.strip() or "in_stock",
            variants=parse_variants(variants),
        )
    except (ValidationError, ValueError) as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc

    async with tenant_session(business.id) as session:
        document = _document_or_404(
            await session.get(KnowledgeDocument, document_id), business.id, KnowledgeDocType.product
        )
        await update_product(document, product)

    return RedirectResponse(url=f"/businesses/{business.id}/dashboard/catalog", status_code=303)


@router.post("/businesses/{business_id}/dashboard/catalog/products/{document_id}/delete")
async def delete_product_route(
    document_id: uuid.UUID, business: Business = Depends(require_business_access)
):
    async with tenant_session(business.id) as session:
        document = _document_or_404(
            await session.get(KnowledgeDocument, document_id), business.id, KnowledgeDocType.product
        )
        await session.delete(document)

    return RedirectResponse(url=f"/businesses/{business.id}/dashboard/catalog", status_code=303)


@router.get("/businesses/{business_id}/dashboard/catalog/promotions/new")
async def new_promotion_form(
    request: Request, business: Business = Depends(require_business_access)
):
    return _render(request, "promotion_form.html", business=business, document=None, promotion=None)


def _promotion_from_form(
    title: str, description: str, discount_text: str, applies_to: str, starts_at: str, ends_at: str
) -> Promotion:
    try:
        return Promotion(
            title=title.strip(),
            description=description.strip(),
            discount_text=discount_text.strip(),
            applies_to=applies_to.strip(),
            starts_at=_parse_amman_datetime(starts_at, field_name="starts_at"),
            ends_at=_parse_amman_datetime(ends_at, field_name="ends_at"),
        )
    except ValidationError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc


@router.post("/businesses/{business_id}/dashboard/catalog/promotions/new")
async def create_promotion_route(
    title: str = Form(...),
    description: str = Form(""),
    discount_text: str = Form(...),
    applies_to: str = Form(""),
    starts_at: str = Form(...),
    ends_at: str = Form(...),
    business: Business = Depends(require_business_access),
):
    promotion = _promotion_from_form(
        title, description, discount_text, applies_to, starts_at, ends_at
    )

    async with tenant_session(business.id) as session:
        await create_promotion(session, business.id, promotion)

    return RedirectResponse(url=f"/businesses/{business.id}/dashboard/catalog", status_code=303)


@router.get("/businesses/{business_id}/dashboard/catalog/promotions/{document_id}/edit")
async def edit_promotion_form(
    request: Request,
    document_id: uuid.UUID,
    business: Business = Depends(require_business_access),
):
    async with tenant_session(business.id) as session:
        document = _document_or_404(
            await session.get(KnowledgeDocument, document_id),
            business.id,
            KnowledgeDocType.promotion,
        )

    return _render(
        request,
        "promotion_form.html",
        business=business,
        document=document,
        promotion=document.structured_data,
        starts_at_local=(
            document.active_from.astimezone(AMMAN_TZ).strftime("%Y-%m-%dT%H:%M")
            if document.active_from
            else ""
        ),
        ends_at_local=(
            document.active_until.astimezone(AMMAN_TZ).strftime("%Y-%m-%dT%H:%M")
            if document.active_until
            else ""
        ),
    )


@router.post("/businesses/{business_id}/dashboard/catalog/promotions/{document_id}/edit")
async def update_promotion_route(
    document_id: uuid.UUID,
    title: str = Form(...),
    description: str = Form(""),
    discount_text: str = Form(...),
    applies_to: str = Form(""),
    starts_at: str = Form(...),
    ends_at: str = Form(...),
    business: Business = Depends(require_business_access),
):
    promotion = _promotion_from_form(
        title, description, discount_text, applies_to, starts_at, ends_at
    )

    async with tenant_session(business.id) as session:
        document = _document_or_404(
            await session.get(KnowledgeDocument, document_id),
            business.id,
            KnowledgeDocType.promotion,
        )
        await update_promotion(document, promotion)

    return RedirectResponse(url=f"/businesses/{business.id}/dashboard/catalog", status_code=303)


@router.post("/businesses/{business_id}/dashboard/catalog/promotions/{document_id}/delete")
async def delete_promotion_route(
    document_id: uuid.UUID, business: Business = Depends(require_business_access)
):
    async with tenant_session(business.id) as session:
        document = _document_or_404(
            await session.get(KnowledgeDocument, document_id),
            business.id,
            KnowledgeDocType.promotion,
        )
        await session.delete(document)

    return RedirectResponse(url=f"/businesses/{business.id}/dashboard/catalog", status_code=303)


@router.post("/businesses/{business_id}/dashboard/approvals/{approval_id}/reject")
async def reject_approval(
    approval_id: uuid.UUID,
    reply_text: str = Form(...),
    business: Business = Depends(require_business_access),
):
    # A <textarea> form submission always normalizes line breaks to \r\n (CRLF) per the HTML
    # spec, but the LLM's own drafted_reply uses plain \n — normalize before comparing, or an
    # unedited approval/resolve would spuriously look "changed" every single time, breaking the
    # adaptive-autonomy streak (graph/nodes/request_owner_approval.py) and falsely recording an
    # owner-correction that's actually the model's own words. Found live during Phase 6d
    # verification: a real dashboard approve, no edits made, still failed `== drafted_reply`.
    reply_text = reply_text.strip().replace("\r\n", "\n")
    if not reply_text:
        raise HTTPException(status_code=400, detail="Reply text is required")

    async with tenant_session(business.id) as session:
        approval = await session.scalar(
            select(ApprovalRequest)
            .where(ApprovalRequest.id == approval_id)
            .options(selectinload(ApprovalRequest.conversation))
        )
        if approval is None or approval.conversation.business_id != business.id:
            raise HTTPException(status_code=404, detail="Approval request not found")
        if approval.status != ApprovalRequestStatus.pending:
            raise HTTPException(status_code=409, detail="Approval request already resolved")

        conversation = approval.conversation

        await send_reply(
            {
                "channel": conversation.channel.value,
                "business_id": str(conversation.business_id),
                "thread_id": conversation.thread_id,
                "draft_reply": {"text": reply_text},
            }
        )

        session.add(
            Message(
                conversation_id=conversation.id,
                direction=MessageDirection.outbound,
                text=reply_text,
                model_used="owner",
            )
        )

        # Declining a same-day commitment isn't a phrasing correction (see plan's Design
        # decisions) — no record_owner_correction call here, unlike the approve route.

        # The Order row estimate_delivery already wrote is still showing the declined same-day
        # promise — update it so it reflects what the customer was actually told, and so it
        # drops out of estimate_delivery's backlog COUNT (which filters on delivery_status ==
        # "pending") for the rest of today.
        if approval.order_reference:
            order = await session.scalar(
                select(Order).where(
                    Order.business_id == business.id,
                    Order.order_reference == approval.order_reference,
                )
            )
            if order is not None:
                order.delivery_status = "declined"
                order.delivery_window_promised = "tomorrow"

        approval.status = ApprovalRequestStatus.rejected
        approval.resolved_by = "owner"
        approval.resolution_time = datetime.now(UTC)
        # A reject is never "unchanged" — set explicitly (rather than left null) so it reads
        # unambiguously alongside sent_unchanged=True/False on approved rows.
        approval.sent_unchanged = False
        conversation.status = ConversationStatus.auto

    return RedirectResponse(url=f"/businesses/{business.id}/dashboard", status_code=303)
