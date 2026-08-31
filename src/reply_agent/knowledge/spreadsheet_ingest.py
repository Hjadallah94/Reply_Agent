"""Parses a seller's catalog workbook (Doc 3 Phase 2: spreadsheet/doc upload -> chunk -> embed)
into the same KnowledgeBase shape the Phase 1 YAML loader produces (knowledge/schema.py) — the
rest of the ingestion pipeline (embedding, storage in knowledge/loader.py) doesn't care which
loader produced it.

Expected workbook: one .xlsx file with up to three sheets, all optional (a seller may not have
FAQs yet), but a present sheet's required columns must all exist:
- "Products": name, price_jod (required); description, stock_status, variants (optional)
    variants cell format: "label:status; label:status", e.g. "size S:in_stock; size L:out_of_stock"
- "Policies": topic, content (both required)
- "FAQs": question, answer (both required)

A missing required column is a hard error for that whole sheet. A bad individual row (wrong
type, empty required cell) is skipped and reported, not fatal to the rest of the sheet.
"""

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from pydantic import ValidationError

from reply_agent.knowledge.schema import FAQPair, KnowledgeBase, Policy, Product, ProductVariant


class WorkbookParseError(ValueError):
    """A sheet is present but missing a required column — nothing usable to ingest from it."""


def _read_rows(sheet: Worksheet) -> tuple[list[str], list[dict[str, object]]]:
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return [], []
    headers = [str(h).strip().lower() if h is not None else "" for h in rows[0]]
    data = [
        {headers[i]: row[i] for i in range(len(headers)) if i < len(row)}
        for row in rows[1:]
        if not all(cell is None for cell in row)
    ]
    return headers, data


def _require_columns(headers: list[str], required: list[str], sheet_name: str) -> None:
    missing = [c for c in required if c not in headers]
    if missing:
        raise WorkbookParseError(
            f"'{sheet_name}' sheet is missing required column(s): {', '.join(missing)}"
        )


def parse_variants(cell: object) -> list[ProductVariant]:
    if not cell or not str(cell).strip():
        return []
    variants = []
    for part in str(cell).split(";"):
        part = part.strip()
        if not part:
            continue
        label, _, status = part.partition(":")
        variants.append(
            ProductVariant(label=label.strip(), stock_status=status.strip() or "in_stock")
        )
    return variants


def parse_catalog_workbook(file, business_slug: str) -> tuple[KnowledgeBase, list[str]]:
    """`file` is anything openpyxl's load_workbook accepts: a path (str/Path) or a
    file-like object such as io.BytesIO."""
    wb = load_workbook(file, read_only=True, data_only=True)
    try:
        issues: list[str] = []

        products: list[Product] = []
        if "Products" in wb.sheetnames:
            headers, rows = _read_rows(wb["Products"])
            _require_columns(headers, ["name", "price_jod"], "Products")
            for i, row in enumerate(rows, start=2):
                try:
                    name = str(row.get("name") or "").strip()
                    if not name:
                        raise ValueError("name is empty")
                    products.append(
                        Product(
                            name=name,
                            description=str(row.get("description") or "").strip(),
                            price_jod=float(row["price_jod"]),
                            stock_status=str(row.get("stock_status") or "in_stock").strip(),
                            variants=parse_variants(row.get("variants")),
                        )
                    )
                except (ValidationError, KeyError, TypeError, ValueError) as exc:
                    issues.append(f"Products row {i}: skipped — {exc}")

        policies: list[Policy] = []
        if "Policies" in wb.sheetnames:
            headers, rows = _read_rows(wb["Policies"])
            _require_columns(headers, ["topic", "content"], "Policies")
            for i, row in enumerate(rows, start=2):
                try:
                    topic = str(row.get("topic") or "").strip()
                    content = str(row.get("content") or "").strip()
                    if not topic or not content:
                        raise ValueError("topic and content are both required")
                    policies.append(Policy(topic=topic, content=content))
                except (ValidationError, KeyError, TypeError, ValueError) as exc:
                    issues.append(f"Policies row {i}: skipped — {exc}")

        faqs: list[FAQPair] = []
        if "FAQs" in wb.sheetnames:
            headers, rows = _read_rows(wb["FAQs"])
            _require_columns(headers, ["question", "answer"], "FAQs")
            for i, row in enumerate(rows, start=2):
                try:
                    question = str(row.get("question") or "").strip()
                    answer = str(row.get("answer") or "").strip()
                    if not question or not answer:
                        raise ValueError("question and answer are both required")
                    faqs.append(FAQPair(question=question, answer=answer))
                except (ValidationError, KeyError, TypeError, ValueError) as exc:
                    issues.append(f"FAQs row {i}: skipped — {exc}")
    finally:
        # read_only workbooks keep their file handle open for lazy streaming until closed —
        # without this, a caller can't delete/reuse the underlying file afterward (fails on
        # Windows with a PermissionError; silently leaks the handle elsewhere).
        wb.close()

    return KnowledgeBase(
        business_slug=business_slug, products=products, policies=policies, faqs=faqs
    ), issues
