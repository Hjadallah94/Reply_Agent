"""Owner dashboard (Doc 3 Phase 3). Real DB; the only thing mocked is the outbound channel
send, same pattern as tests/unit/test_send_reply.py. Every route here is gated by
auth/dependencies.py — tests log in via tests/auth_helpers.py the same way a browser would.
"""

import uuid
from datetime import UTC, datetime
from io import BytesIO
from unittest.mock import AsyncMock, patch

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook
from sqlalchemy import delete, select

from reply_agent.api.app import app
from reply_agent.db.models import (
    ApprovalRequest,
    ApprovalRequestStatus,
    Business,
    ChannelType,
    Conversation,
    ConversationStatus,
    Customer,
    Escalation,
    EscalationStatus,
    KnowledgeDocType,
    KnowledgeDocument,
    Message,
    MessageDirection,
    Order,
)
from reply_agent.db.session import get_sessionmaker
from reply_agent.graph.nodes.request_owner_approval import AUTO_APPROVAL_RESOLVED_BY
from tests.auth_helpers import create_logged_in_business, dispose_engines

BUSINESS_NAME = "Dashboard Test Business"


@pytest.fixture
def client():
    return TestClient(app)


@pytest.fixture
async def escalation(client):
    business = await create_logged_in_business(client, BUSINESS_NAME)

    async with get_sessionmaker()() as session:
        db_business = await session.get(Business, business.id)
        db_business.channels_connected = {"whatsapp": {"phone_number_id": "test-phone-number-id"}}

        customer = Customer(
            business_id=business.id,
            channel=ChannelType.whatsapp,
            channel_handle="962790001111",
        )
        session.add(customer)
        await session.flush()

        conversation = Conversation(
            business_id=business.id,
            channel=ChannelType.whatsapp,
            customer_id=customer.id,
            status=ConversationStatus.owner_handled,
            thread_id=f"whatsapp:{business.id}:962790001111",
        )
        session.add(conversation)
        await session.flush()

        session.add(
            Message(
                conversation_id=conversation.id,
                direction=MessageDirection.inbound,
                text="Can I get a refund?",
            )
        )

        escalation = Escalation(
            conversation_id=conversation.id,
            reason="refund_or_complaint always escalates",
            drafted_reply="Sorry, no cash refunds — exchanges only.",
            status=EscalationStatus.pending,
        )
        session.add(escalation)
        await session.commit()
        await session.refresh(escalation)

    # About to hand control to the test body, which makes its own TestClient calls — same
    # cross-loop issue as create_logged_in_business's own dispose, needed again here since
    # this fixture did more direct session work afterward.
    await dispose_engines()
    yield business, escalation

    # The test body made its own TestClient calls after this session block closed — same
    # cross-loop issue create_logged_in_business's own dispose handles, needed again here
    # before this fixture's own teardown DB access.
    await dispose_engines()
    async with get_sessionmaker()() as session:
        await session.execute(delete(Business).where(Business.id == business.id))
        await session.commit()


@pytest.fixture
async def approval(client):
    business = await create_logged_in_business(client, "Approval " + BUSINESS_NAME)

    async with get_sessionmaker()() as session:
        db_business = await session.get(Business, business.id)
        db_business.channels_connected = {"whatsapp": {"phone_number_id": "test-phone-number-id"}}

        customer = Customer(
            business_id=business.id,
            channel=ChannelType.whatsapp,
            channel_handle="962790002222",
        )
        session.add(customer)
        await session.flush()

        conversation = Conversation(
            business_id=business.id,
            channel=ChannelType.whatsapp,
            customer_id=customer.id,
            status=ConversationStatus.owner_handled,
            thread_id=f"whatsapp:{business.id}:962790002222",
        )
        session.add(conversation)
        await session.flush()

        session.add(
            Message(
                conversation_id=conversation.id,
                direction=MessageDirection.inbound,
                text="2 boxes of chocolate chip to Sweifieh please",
            )
        )

        order = Order(
            business_id=business.id,
            order_reference="chat-test1234",
            customer_phone="962790002222",
            status="pending_delivery_estimate",
            items_summary="2 item(s)",
            delivery_address="Sweifieh, Amman",
            delivery_window_promised="3-4 hours",
            delivery_status="pending",
        )
        session.add(order)

        approval = ApprovalRequest(
            conversation_id=conversation.id,
            drafted_reply="We can get that to you in 3-4 hours today.",
            reasoning="0 order(s) already pending today, ~45 min transit time.",
            order_reference="chat-test1234",
            status=ApprovalRequestStatus.pending,
        )
        session.add(approval)
        await session.commit()
        await session.refresh(approval)

    await dispose_engines()
    yield business, approval

    await dispose_engines()
    async with get_sessionmaker()() as session:
        await session.execute(delete(Business).where(Business.id == business.id))
        await session.commit()


async def test_dashboard_redirects_to_the_logged_in_user_own_business(client, escalation):
    business, _ = escalation
    response = client.get("/dashboard", follow_redirects=False)
    assert response.status_code == 303
    assert response.headers["location"] == f"/businesses/{business.id}/dashboard"


async def test_dashboard_redirects_to_login_when_not_authenticated():
    client = TestClient(app)
    response = client.get("/dashboard", follow_redirects=False)
    assert response.status_code == 303
    assert response.headers["location"] == "/login"


# --- Bilingual dashboard (Doc 3 Phase 6.6) -------------------------------------------------


async def test_dashboard_defaults_to_english_and_ltr(client, escalation):
    business, _ = escalation
    response = client.get(f"/businesses/{business.id}/dashboard")
    assert response.status_code == 200
    assert 'dir="ltr"' in response.text
    assert "Needs your reply" in response.text


async def test_set_language_to_arabic_persists_and_flips_rtl(client, escalation):
    business, _ = escalation
    response = client.post(
        "/set-language",
        data={"lang": "ar", "next": f"/businesses/{business.id}/dashboard"},
        follow_redirects=False,
    )
    assert response.status_code == 303
    assert response.headers["location"] == f"/businesses/{business.id}/dashboard"

    # Each TestClient call spins its own event loop (see test_resolve_already_resolved_returns_409
    # above) — dispose between calls, not just around raw DB session access.
    await dispose_engines()
    dashboard_response = client.get(f"/businesses/{business.id}/dashboard")
    assert dashboard_response.status_code == 200
    assert 'dir="rtl"' in dashboard_response.text
    assert "بحاجة لردك" in dashboard_response.text  # "Needs your reply"


async def test_set_language_rejects_unsupported_language(client, escalation):
    response = client.post("/set-language", data={"lang": "fr", "next": "/dashboard"})
    assert response.status_code == 400


async def test_set_language_guards_against_open_redirect(client, escalation):
    response = client.post(
        "/set-language",
        data={"lang": "ar", "next": "https://evil.example.com/phish"},
        follow_redirects=False,
    )
    assert response.status_code == 303
    assert response.headers["location"] == "/dashboard"


async def test_business_dashboard_lists_the_escalation(client, escalation):
    business, esc = escalation
    response = client.get(f"/businesses/{business.id}/dashboard")
    assert response.status_code == 200
    assert "962790001111" in response.text
    assert f"/dashboard/escalations/{esc.id}" in response.text


async def test_business_dashboard_404s_for_unknown_business(client, escalation):
    # Logged in, but to a different business than the one being requested.
    response = client.get("/businesses/00000000-0000-0000-0000-000000000000/dashboard")
    assert response.status_code == 404


async def test_business_dashboard_redirects_to_login_when_not_authenticated(escalation):
    business, _ = escalation
    anonymous_client = TestClient(app)
    response = anonymous_client.get(f"/businesses/{business.id}/dashboard", follow_redirects=False)
    assert response.status_code == 303
    assert response.headers["location"] == "/login"


async def test_escalation_detail_shows_draft_and_reason(client, escalation):
    business, esc = escalation
    response = client.get(f"/businesses/{business.id}/dashboard/escalations/{esc.id}")
    assert response.status_code == 200
    assert "refund_or_complaint always escalates" in response.text
    assert "Sorry, no cash refunds" in response.text
    assert "Can I get a refund?" in response.text


async def test_resolve_sends_updates_db_and_redirects(client, escalation):
    business, esc = escalation

    with (
        patch(
            "reply_agent.graph.nodes.send_reply.send_whatsapp_message", new=AsyncMock()
        ) as mock_send,
        patch(
            "reply_agent.knowledge.corrections.embed_documents",
            return_value=[[0.0] * 1024],
        ),
    ):
        response = client.post(
            f"/businesses/{business.id}/dashboard/escalations/{esc.id}/resolve",
            data={"reply_text": "Edited final reply"},
            follow_redirects=False,
        )

    assert response.status_code == 303
    assert response.headers["location"] == f"/businesses/{business.id}/dashboard"
    mock_send.assert_called_once_with(
        to="962790001111", text="Edited final reply", phone_number_id="test-phone-number-id"
    )

    await dispose_engines()

    async with get_sessionmaker()() as session:
        refreshed = await session.get(Escalation, esc.id)
        assert refreshed.status == EscalationStatus.resolved
        assert refreshed.resolved_by == "owner"
        assert refreshed.resolution_text == "Edited final reply"

        conversation = await session.get(Conversation, esc.conversation_id)
        assert conversation.status == ConversationStatus.auto

        outbound = await session.scalar(
            select(Message).where(
                Message.conversation_id == esc.conversation_id,
                Message.direction == MessageDirection.outbound,
            )
        )
        assert outbound.text == "Edited final reply"
        assert outbound.model_used == "owner"


async def test_resolve_with_an_edited_reply_records_a_correction(client, escalation):
    """Owner-correction feedback loop (Doc 1 Section 7) — sending something different from the
    agent's own draft should be captured as a new brand_voice few-shot example.
    """
    business, esc = escalation

    with (
        patch("reply_agent.graph.nodes.send_reply.send_whatsapp_message", new=AsyncMock()),
        patch(
            "reply_agent.knowledge.corrections.embed_documents",
            return_value=[[0.0] * 1024],
        ) as mock_embed,
    ):
        client.post(
            f"/businesses/{business.id}/dashboard/escalations/{esc.id}/resolve",
            data={"reply_text": "Edited final reply"},
            follow_redirects=False,
        )

    mock_embed.assert_called_once()
    embedded_text = mock_embed.call_args[0][0][0]
    assert "Can I get a refund?" in embedded_text
    assert "Edited final reply" in embedded_text

    await dispose_engines()

    async with get_sessionmaker()() as session:
        correction = await session.scalar(
            select(KnowledgeDocument).where(
                KnowledgeDocument.business_id == business.id,
                KnowledgeDocument.type == KnowledgeDocType.brand_voice,
            )
        )
        assert correction is not None
        assert correction.structured_data["source"] == "owner_correction"
        assert correction.structured_data["escalation_id"] == str(esc.id)


async def test_resolve_approving_the_draft_unchanged_records_no_correction(client, escalation):
    business, esc = escalation

    with (
        patch("reply_agent.graph.nodes.send_reply.send_whatsapp_message", new=AsyncMock()),
        patch("reply_agent.knowledge.corrections.embed_documents") as mock_embed,
    ):
        client.post(
            f"/businesses/{business.id}/dashboard/escalations/{esc.id}/resolve",
            data={"reply_text": esc.drafted_reply},
            follow_redirects=False,
        )

    mock_embed.assert_not_called()

    await dispose_engines()

    async with get_sessionmaker()() as session:
        correction = await session.scalar(
            select(KnowledgeDocument).where(
                KnowledgeDocument.business_id == business.id,
                KnowledgeDocument.type == KnowledgeDocType.brand_voice,
            )
        )
        assert correction is None


async def test_resolve_unchanged_with_crlf_line_endings_still_counts_as_unchanged(
    client, escalation
):
    """Same bug as the approval-route regression test above — a <textarea> submission
    normalizes to \\r\\n while drafted_reply uses plain \\n; an unedited resolve must not be
    misrecorded as an owner correction just because of line-ending noise.
    """
    business, esc = escalation

    async with get_sessionmaker()() as session:
        db_escalation = await session.get(Escalation, esc.id)
        db_escalation.drafted_reply = "Line one.\nLine two."
        await session.commit()
    await dispose_engines()

    with (
        patch("reply_agent.graph.nodes.send_reply.send_whatsapp_message", new=AsyncMock()),
        patch("reply_agent.knowledge.corrections.embed_documents") as mock_embed,
    ):
        client.post(
            f"/businesses/{business.id}/dashboard/escalations/{esc.id}/resolve",
            data={"reply_text": "Line one.\r\nLine two."},
            follow_redirects=False,
        )

    mock_embed.assert_not_called()


async def test_resolve_already_resolved_returns_409(client, escalation):
    business, esc = escalation

    # Set up the "already resolved" state directly via the ORM rather than a prior HTTP call —
    # two TestClient calls in one test each spin their own event loop, and the connection pool
    # from the first breaks when the second reuses it (deeper than the dispose-between-calls
    # workaround used elsewhere in this file covers).
    async with get_sessionmaker()() as session:
        db_escalation = await session.get(Escalation, esc.id)
        db_escalation.status = EscalationStatus.resolved
        await session.commit()
    await dispose_engines()

    with patch("reply_agent.graph.nodes.send_reply.send_whatsapp_message", new=AsyncMock()):
        response = client.post(
            f"/businesses/{business.id}/dashboard/escalations/{esc.id}/resolve",
            data={"reply_text": "Second attempt"},
        )

    assert response.status_code == 409


async def test_resolve_rejects_blank_reply(client, escalation):
    business, esc = escalation
    response = client.post(
        f"/businesses/{business.id}/dashboard/escalations/{esc.id}/resolve",
        data={"reply_text": "   "},
    )
    assert response.status_code == 400


async def test_export_returns_xlsx_with_messages(client, escalation):
    business, _esc = escalation
    response = client.get(f"/businesses/{business.id}/dashboard/export")

    assert response.status_code == 200
    assert response.headers["content-type"] == (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert "attachment" in response.headers["content-disposition"]

    workbook = load_workbook(BytesIO(response.content))
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    assert rows[0] == (
        "Customer",
        "Channel",
        "From",
        "Message",
        "Intent",
        "Handled by",
        "Conversation status",
        "Sent at (UTC)",
    )
    assert rows[1][0] == "962790001111"
    assert rows[1][2] == "Customer"
    assert rows[1][3] == "Can I get a refund?"


async def test_export_404s_for_unknown_business(client, escalation):
    response = client.get("/businesses/00000000-0000-0000-0000-000000000000/dashboard/export")
    assert response.status_code == 404


async def test_business_dashboard_lists_the_approval(client, approval):
    business, appr = approval
    response = client.get(f"/businesses/{business.id}/dashboard")
    assert response.status_code == 200
    assert "962790002222" in response.text
    assert f"/dashboard/approvals/{appr.id}" in response.text


async def test_approval_detail_shows_draft_and_reasoning(client, approval):
    business, appr = approval
    response = client.get(f"/businesses/{business.id}/dashboard/approvals/{appr.id}")
    assert response.status_code == 200
    assert "0 order(s) already pending today" in response.text
    assert "We can get that to you in 3-4 hours today." in response.text
    assert "2 boxes of chocolate chip to Sweifieh please" in response.text


async def test_approve_sends_updates_db_and_redirects(client, approval):
    business, appr = approval

    with (
        patch(
            "reply_agent.graph.nodes.send_reply.send_whatsapp_message", new=AsyncMock()
        ) as mock_send,
        patch(
            "reply_agent.knowledge.corrections.embed_documents",
            return_value=[[0.0] * 1024],
        ),
    ):
        response = client.post(
            f"/businesses/{business.id}/dashboard/approvals/{appr.id}/approve",
            data={"reply_text": "Edited: 3 hours today"},
            follow_redirects=False,
        )

    assert response.status_code == 303
    assert response.headers["location"] == f"/businesses/{business.id}/dashboard"
    mock_send.assert_called_once_with(
        to="962790002222", text="Edited: 3 hours today", phone_number_id="test-phone-number-id"
    )

    await dispose_engines()

    async with get_sessionmaker()() as session:
        refreshed = await session.get(ApprovalRequest, appr.id)
        assert refreshed.status == ApprovalRequestStatus.approved
        assert refreshed.resolved_by == "owner"
        # Edited before sending — doesn't count toward the adaptive-autonomy streak
        # (graph/nodes/request_owner_approval.py's _matches_learned_pattern).
        assert refreshed.sent_unchanged is False

        conversation = await session.get(Conversation, appr.conversation_id)
        assert conversation.status == ConversationStatus.auto

        outbound = await session.scalar(
            select(Message).where(
                Message.conversation_id == appr.conversation_id,
                Message.direction == MessageDirection.outbound,
            )
        )
        assert outbound.text == "Edited: 3 hours today"
        assert outbound.model_used == "owner"

        # Approving doesn't touch the Order row — the same-day promise it already holds
        # is exactly what got approved.
        order = await session.scalar(select(Order).where(Order.order_reference == "chat-test1234"))
        assert order.delivery_status == "pending"
        assert order.delivery_window_promised == "3-4 hours"


async def test_approve_with_an_edited_reply_records_a_correction(client, approval):
    business, appr = approval

    with (
        patch("reply_agent.graph.nodes.send_reply.send_whatsapp_message", new=AsyncMock()),
        patch(
            "reply_agent.knowledge.corrections.embed_documents",
            return_value=[[0.0] * 1024],
        ) as mock_embed,
    ):
        client.post(
            f"/businesses/{business.id}/dashboard/approvals/{appr.id}/approve",
            data={"reply_text": "Edited: 3 hours today"},
            follow_redirects=False,
        )

    mock_embed.assert_called_once()

    await dispose_engines()

    async with get_sessionmaker()() as session:
        correction = await session.scalar(
            select(KnowledgeDocument).where(
                KnowledgeDocument.business_id == business.id,
                KnowledgeDocument.type == KnowledgeDocType.brand_voice,
            )
        )
        assert correction is not None
        assert correction.structured_data["source"] == "owner_correction"
        assert correction.structured_data["approval_id"] == str(appr.id)


async def test_approve_unchanged_records_no_correction(client, approval):
    business, appr = approval

    with (
        patch("reply_agent.graph.nodes.send_reply.send_whatsapp_message", new=AsyncMock()),
        patch("reply_agent.knowledge.corrections.embed_documents") as mock_embed,
    ):
        client.post(
            f"/businesses/{business.id}/dashboard/approvals/{appr.id}/approve",
            data={"reply_text": appr.drafted_reply},
            follow_redirects=False,
        )

    mock_embed.assert_not_called()

    await dispose_engines()

    async with get_sessionmaker()() as session:
        refreshed = await session.get(ApprovalRequest, appr.id)
        assert refreshed.sent_unchanged is True


async def test_approve_unchanged_with_crlf_line_endings_still_counts_as_unchanged(
    client, approval
):
    """A <textarea> form submission always normalizes line breaks to \\r\\n per the HTML spec,
    but drafted_reply (the LLM's own output) uses plain \\n — a real bug found live during
    Phase 6d verification: an unedited approval submitted through an actual browser was
    incorrectly recorded as sent_unchanged=False, silently breaking the adaptive-autonomy
    streak every single time.
    """
    business, appr = approval

    async with get_sessionmaker()() as session:
        db_approval = await session.get(ApprovalRequest, appr.id)
        db_approval.drafted_reply = "Line one.\nLine two."
        await session.commit()
    await dispose_engines()

    with (
        patch("reply_agent.graph.nodes.send_reply.send_whatsapp_message", new=AsyncMock()),
        patch("reply_agent.knowledge.corrections.embed_documents") as mock_embed,
    ):
        client.post(
            f"/businesses/{business.id}/dashboard/approvals/{appr.id}/approve",
            data={"reply_text": "Line one.\r\nLine two."},
            follow_redirects=False,
        )

    mock_embed.assert_not_called()

    await dispose_engines()
    async with get_sessionmaker()() as session:
        refreshed = await session.get(ApprovalRequest, appr.id)
        assert refreshed.sent_unchanged is True


async def test_reject_sends_tomorrow_message_and_updates_order(client, approval):
    business, appr = approval

    with patch(
        "reply_agent.graph.nodes.send_reply.send_whatsapp_message", new=AsyncMock()
    ) as mock_send:
        response = client.post(
            f"/businesses/{business.id}/dashboard/approvals/{appr.id}/reject",
            data={"reply_text": "Sorry, not approved — delivery is tomorrow instead."},
            follow_redirects=False,
        )

    assert response.status_code == 303
    mock_send.assert_called_once_with(
        to="962790002222",
        text="Sorry, not approved — delivery is tomorrow instead.",
        phone_number_id="test-phone-number-id",
    )

    await dispose_engines()

    async with get_sessionmaker()() as session:
        refreshed = await session.get(ApprovalRequest, appr.id)
        assert refreshed.status == ApprovalRequestStatus.rejected
        assert refreshed.sent_unchanged is False

        conversation = await session.get(Conversation, appr.conversation_id)
        assert conversation.status == ConversationStatus.auto

        # The Order row estimate_delivery wrote must no longer show the declined same-day
        # promise — this is the part that would otherwise be silently wrong.
        order = await session.scalar(select(Order).where(Order.order_reference == "chat-test1234"))
        assert order.delivery_status == "declined"
        assert order.delivery_window_promised == "tomorrow"

        correction = await session.scalar(
            select(KnowledgeDocument).where(
                KnowledgeDocument.business_id == business.id,
                KnowledgeDocument.type == KnowledgeDocType.brand_voice,
            )
        )
        assert correction is None


async def test_approval_already_resolved_returns_409(client, approval):
    business, appr = approval

    async with get_sessionmaker()() as session:
        db_approval = await session.get(ApprovalRequest, appr.id)
        db_approval.status = ApprovalRequestStatus.approved
        await session.commit()
    await dispose_engines()

    with patch("reply_agent.graph.nodes.send_reply.send_whatsapp_message", new=AsyncMock()):
        response = client.post(
            f"/businesses/{business.id}/dashboard/approvals/{appr.id}/approve",
            data={"reply_text": "Second attempt"},
        )

    assert response.status_code == 409


async def test_approval_rejects_blank_reply(client, approval):
    business, appr = approval
    response = client.post(
        f"/businesses/{business.id}/dashboard/approvals/{appr.id}/reject",
        data={"reply_text": "   "},
    )
    assert response.status_code == 400


async def test_dashboard_shows_recently_auto_approved_section(client, approval):
    business, appr = approval

    async with get_sessionmaker()() as session:
        session.add(
            ApprovalRequest(
                id=uuid.uuid4(),
                conversation_id=appr.conversation_id,
                drafted_reply="Auto-sent draft.",
                reasoning="Learned pattern.",
                estimated_window="3-4 hours",
                status=ApprovalRequestStatus.approved,
                resolved_by=AUTO_APPROVAL_RESOLVED_BY,
                sent_unchanged=True,
                resolution_time=datetime.now(UTC),
            )
        )
        await session.commit()
    await dispose_engines()

    response = client.get(f"/businesses/{business.id}/dashboard")
    assert response.status_code == 200
    assert "Recently auto-approved" in response.text
    assert "962790002222" in response.text
    assert "3-4 hours" in response.text


async def test_dashboard_omits_auto_approved_section_when_theres_none(client, approval):
    business, _appr = approval
    response = client.get(f"/businesses/{business.id}/dashboard")
    assert response.status_code == 200
    assert "Recently auto-approved" not in response.text
