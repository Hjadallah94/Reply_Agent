import io

import pytest
from openpyxl import Workbook

from reply_agent.orders.spreadsheet_ingest import OrderWorkbookParseError, parse_orders_workbook


def _workbook(rows: list[list]) -> io.BytesIO:
    wb = Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("Orders")
    for row in rows:
        ws.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def test_parses_valid_orders():
    wb = _workbook(
        [
            [
                "order_reference",
                "customer_phone",
                "customer_name",
                "status",
                "items_summary",
                "order_date",
            ],
            ["ORD-1", "0791234567", "Sara", "shipped", "Black Abaya x1", "2026-08-20"],
        ]
    )

    records, issues = parse_orders_workbook(wb)

    assert issues == []
    assert len(records) == 1
    record = records[0]
    assert record.order_reference == "ORD-1"
    assert record.customer_phone == "962791234567"
    assert record.status == "shipped"
    assert record.order_date.isoformat().startswith("2026-08-20")


def test_missing_sheet_raises():
    wb = Workbook()
    wb.remove(wb.active)
    wb.create_sheet("NotOrders")
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    with pytest.raises(OrderWorkbookParseError, match="Orders"):
        parse_orders_workbook(buf)


def test_missing_required_column_raises():
    wb = _workbook([["order_reference", "customer_phone"], ["ORD-1", "0791234567"]])

    with pytest.raises(OrderWorkbookParseError, match="status"):
        parse_orders_workbook(wb)


def test_bad_row_is_skipped_not_fatal():
    wb = _workbook(
        [
            ["order_reference", "customer_phone", "status"],
            ["ORD-1", "0791234567", "shipped"],
            ["ORD-2", "", "shipped"],
        ]
    )

    records, issues = parse_orders_workbook(wb)

    assert len(records) == 1
    assert records[0].order_reference == "ORD-1"
    assert len(issues) == 1
    assert "row 3" in issues[0]


def test_optional_fields_default_to_none():
    wb = _workbook(
        [["order_reference", "customer_phone", "status"], ["ORD-1", "0791234567", "shipped"]]
    )

    records, _ = parse_orders_workbook(wb)

    assert records[0].customer_name is None
    assert records[0].items_summary is None
    assert records[0].order_date is None
